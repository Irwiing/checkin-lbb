from pathlib import Path

from flask import Flask, redirect, render_template, request, url_for
from openpyxl import Workbook, load_workbook
import qrcode

app = Flask(__name__)

BASE_DIR = Path(__file__).resolve().parent
QR_CODE_PATH = BASE_DIR / "static" / "qrcode.png"
XLS_PATH = BASE_DIR / "nomes.xlsx"


def ensure_workbook() -> None:
    if XLS_PATH.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nomes"
    sheet.append(["Nome"])
    workbook.save(XLS_PATH)


def save_name(name: str) -> None:
    ensure_workbook()

    workbook = load_workbook(XLS_PATH)
    sheet = workbook.active
    sheet.append([name.strip()])
    workbook.save(XLS_PATH)


def generate_qr() -> None:
    QR_CODE_PATH.parent.mkdir(parents=True, exist_ok=True)

    if QR_CODE_PATH.exists():
        return

    url = "http://localhost:5000/form"
    image = qrcode.make(url)
    image.save(QR_CODE_PATH)


@app.route("/")
def index():
    generate_qr()
    return render_template("index.html")


@app.route("/form", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        name = request.form.get("name", "")
        if name.strip():
            save_name(name)
            return redirect(url_for("success"))
    return render_template("form.html")


@app.route("/success")
def success():
    return render_template("success.html")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
